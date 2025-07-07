import pandas as pd
import warnings
warnings.filterwarnings('ignore')

# Read providers.xlsx
df_providers = pd.read_excel('docs/requirements_material/providers.xlsx')

# Get detailed info for JCL, BOOST, and CHINHIN
print('=== DETAILED PROVIDER INFORMATION ===\n')

target_providers = ['JCL', 'BOOST', 'CHINHIN']

for provider in target_providers:
    print(f'\n{provider}:')
    row = df_providers[df_providers['Short  name'] == provider]
    if not row.empty:
        row = row.iloc[0]
        print(f'  - ID: {row["id"]}')
        print(f'  - Reference Company ID: {row["ref_comid"]}')
        print(f'  - Size Group: {row["Size Group"]}')
        print(f'  - History Period: {row["From"]} to {row["To"]}')
        print(f'  - From Date: {row["From (Date)"]}')
        print(f'  - To Date: {row["To (Date)"]}')
        print(f'  - Recency Lag: {row["Recency Lag"]}')
        
        # Check recent monthly data
        recent_months = ['2024-10', '2024-11', '2024-12', '2025-01', '2025-02', '2025-03', '2025-04', '2025-05']
        print(f'  - Recent monthly values:')
        for month in recent_months:
            if month in df_providers.columns:
                value = row[month]
                if pd.notna(value) and value != 0:
                    print(f'    {month}: {value:,.0f}')

# Check for additional sheets
print('\n\n=== CHECKING FOR ADDITIONAL SHEETS ===')
import openpyxl

# Check draft_dictionary.xlsx sheets
wb1 = openpyxl.load_workbook('docs/dictionaries/draft_dictionary.xlsx', read_only=True)
print(f'\nSheets in draft_dictionary.xlsx: {wb1.sheetnames}')
wb1.close()

# Check providers.xlsx sheets  
wb2 = openpyxl.load_workbook('docs/requirements_material/providers.xlsx', read_only=True)
print(f'\nSheets in providers.xlsx: {wb2.sheetnames}')
wb2.close()